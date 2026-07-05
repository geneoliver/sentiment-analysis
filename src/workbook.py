"""Assembles the final 5-tab .xlsx workbook.

Tabs 3-5 are flat pivot-style rollups by design (payee/category, total,
count, date range) — no merged cells or nesting, so they stay easy to
re-derive from tab 2 if you ever need to.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

RAW_DATA_SHEET = "Raw Data"
CATEGORIZED_SHEET = "Categorized Transactions"
BY_PAYEE_SHEET = "Summary by Payee"
BY_CATEGORY_SHEET = "Summary by Category"
BY_ESSENTIAL_DISCRETIONARY_SHEET = "Essential vs Discretionary"

CURRENCY_COLUMNS = {"amount", "total_amount"}


def build_workbook(
    output_path: Path,
    raw_df: pd.DataFrame,
    categorized_df: pd.DataFrame,
    by_payee_df: pd.DataFrame,
    by_category_df: pd.DataFrame,
    by_essential_discretionary_df: pd.DataFrame,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sheets = {
        RAW_DATA_SHEET: raw_df,
        CATEGORIZED_SHEET: categorized_df,
        BY_PAYEE_SHEET: by_payee_df,
        BY_CATEGORY_SHEET: by_category_df,
        BY_ESSENTIAL_DISCRETIONARY_SHEET: by_essential_discretionary_df,
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)

        for name, df in sheets.items():
            _format_sheet(writer.sheets[name], df)


def _format_sheet(worksheet, df: pd.DataFrame) -> None:
    bold = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = bold

    for col_idx, column in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = max([len(str(column))] + [len(str(v)) for v in df[column].head(200)]) if len(df) else len(str(column))
        worksheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

        if column in CURRENCY_COLUMNS:
            for row_idx in range(2, len(df) + 2):
                worksheet.cell(row=row_idx, column=col_idx).number_format = '$#,##0.00'

    worksheet.freeze_panes = "A2"
